from pathlib import Path
from io import BytesIO
import os

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.routers.choke_costing_ui_router import _unique_upload_path
from services.customer_input_extraction import (
    apply_resolution_to_customer_input,
    discover_customer_input_files,
    extract_customer_input_package,
    extract_excel_fields,
    merge_customer_input_candidates,
    normalize_drawing_reference,
    normalize_product_family,
    normalize_delivery_location,
)
from services.customer_input_schema import normalize_customer_input


ACTUAL_OLIVIER_WORKBOOK = Path(os.getenv(
    "OLIVIER_GREEN_WORKBOOK_PATH",
    (
        r"C:\Users\youssef.benamor\Downloads\OneDrive_2026-07-22"
        r"\01- Customer Input\Choke2024-06-22T06_54_40.0797274Z.xlsm"
    ),
))


def _workbook(path: Path, *, hidden_currency=None, layout="rows") -> Path:
    workbook = openpyxl.Workbook()
    try:
        sheet = workbook.active
        sheet.title = "RFQ"
        values = {
            "Project code": "24018-CHO-00",
            "Customer": "PRABHA ENGINEERING",
            "Final customer": "Prabha Engineering",
            "Product": "Rod Choke",
            "Product line": "Choke",
            "Part number": "300440157",
            "SOP": "2024-12-01",
            "Delivery country": "India",
            "Delivery city": "Pune",
            "Quotation currency": "INR",
        }
        if layout == "rows":
            for row, (label, value) in enumerate(values.items(), 1):
                sheet.cell(row, 1, label)
                sheet.cell(row, 2, value)
        else:
            for column, (label, value) in enumerate(values.items(), 1):
                sheet.cell(1, column, label)
                sheet.cell(2, column, value)
                sheet.cell(1, column).font = openpyxl.styles.Font(bold=True)
        year_row = 20
        sheet.cell(year_row - 1, 1, "Annual volume")
        for column, year in enumerate(range(2024, 2029), 2):
            sheet.cell(year_row, column, year)
            sheet.cell(year_row + 1, column, 360000)
        hidden = workbook.create_sheet("Finance")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "Quotation currency"
        hidden["B1"] = hidden_currency or "USD"
        workbook.create_named_range("BrokenCurrency", sheet, "#REF!")
        workbook.save(path)
    finally:
        workbook.close()
    return path


@pytest.mark.skipif(
    not ACTUAL_OLIVIER_WORKBOOK.exists(),
    reason="Actual Olivier RFQ workbook is not available on this machine.",
)
def test_actual_olivier_workbook_extraction_has_cell_provenance():
    report = extract_excel_fields(ACTUAL_OLIVIER_WORKBOOK)
    by_field = {}
    for candidate in report["candidates"]:
        by_field.setdefault(candidate["field"], []).append(candidate)
    assert any(
        item["value"] == "Rod Choke" and item["source_cell"] == "K4"
        for item in by_field["product_name"]
    )
    assert any(
        str(item["value"]) == "300440157" and item["source_cell"] == "N4"
        for item in by_field["part_number"]
    )
    assert any(
        item["value"] == "INR" and item["source_cell"] == "R14"
        for item in by_field["quotation_currency"]
    )
    assert report["green_commercial_inputs"] == []
    assert report["unmapped_green_inputs"] == []


def _pdf(path: Path) -> Path:
    path.write_bytes(b"%PDF-1.4\n%%EOF")
    return path


def test_pdf_plus_xlsm_resolves_all_mandatory_fields(tmp_path):
    workbook = _workbook(tmp_path / "input.xlsm")
    pdf = _pdf(tmp_path / "drawing.pdf")
    result = extract_customer_input_package({}, [workbook, pdf])
    assert result["component_costing_ready"] is True
    assert result["missing_fields"] == []


def test_workbook_commercial_fields_merge_with_pdf_metadata(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "input.xlsx"), _pdf(tmp_path / "drawing.pdf")])
    assert result["fields"]["quotation_currency"]["value"] == "INR"
    assert result["fields"]["drawing_reference"]["value"] == "drawing.pdf"


def test_yearly_series_uses_max_without_summing(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "input.xlsx")])
    annual = result["fields"]["annual_quantity"]
    assert annual["value"] == 360000
    assert annual["qmax"] == 360000
    assert len(annual["quantity_by_year"]) == 5


def test_broken_names_are_ignored_with_warning(tmp_path):
    report = extract_excel_fields(_workbook(tmp_path / "input.xlsx"))
    assert any("broken or unavailable named range" in item for item in report["warnings"])


def test_hidden_financial_sheet_cannot_override_visible_value(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "input.xlsx", hidden_currency="USD")])
    assert result["fields"]["quotation_currency"]["value"] == "INR"


def test_explicit_payload_overrides_workbook_without_blocking(tmp_path):
    payload = {"quotation_currency": "EUR", "_explicit_user_fields": ["quotation_currency"]}
    result = extract_customer_input_package(payload, [_workbook(tmp_path / "input.xlsx")])
    assert result["fields"]["quotation_currency"]["value"] == "EUR"
    assert not any(item["field"] == "quotation_currency" for item in result["conflicts"])


def test_unconfirmed_structured_payload_conflict_blocks(tmp_path):
    result = extract_customer_input_package({"quotation_currency": "EUR"}, [_workbook(tmp_path / "input.xlsx")])
    conflict = next(item for item in result["conflicts"] if item["field"] == "quotation_currency")
    assert conflict["blocking"] is True
    assert result["component_costing_ready"] is False


def test_supplier_currency_does_not_satisfy_quotation_currency():
    result = extract_customer_input_package({"supplier_currency": "CNY"}, [])
    assert "quotation_currency" in result["missing_fields"]


def test_country_normalizes_to_existing_zone():
    result = normalize_delivery_location(country="China", city="Kunshan")
    assert result["normalized_zone"].lower() == "china south pacific"
    assert result["mapping_source"] == "country_to_avocarbon_zone"


def test_repeated_extraction_is_idempotent(tmp_path):
    files = [_workbook(tmp_path / "input.xlsx"), _pdf(tmp_path / "drawing.pdf")]
    first = extract_customer_input_package({}, files)
    second = extract_customer_input_package({}, files)
    assert first["fields"] == second["fields"]
    assert first["conflicts"] == second["conflicts"]


def test_duplicate_filenames_receive_unique_paths(tmp_path):
    first = tmp_path / "drawing.pdf"
    first.write_bytes(b"one")
    second = _unique_upload_path(tmp_path, "drawing.pdf")
    assert second.name == "drawing__2.pdf"


def test_pdf_only_flow_remains_supported(tmp_path):
    pdf = _pdf(tmp_path / "drawing.pdf")
    discovered = discover_customer_input_files([pdf])
    result = extract_customer_input_package({}, discovered)
    assert result["fields"]["drawing_reference"]["value"] == "drawing.pdf"
    assert "annual_quantity" in result["missing_fields"]


def test_different_workbook_layout_is_parsed_by_labels(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "columns.xlsx", layout="columns")])
    assert result["fields"]["customer"]["value"] == "PRABHA ENGINEERING"
    assert result["fields"]["part_number"]["value"] == "300440157"


def test_production_plant_is_not_inferred_from_delivery_location(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "input.xlsx")])
    customer_input = apply_resolution_to_customer_input({}, result)
    assert customer_input.get("production_plant") is None


def test_target_price_is_optional_for_component_costing(tmp_path):
    result = extract_customer_input_package({}, [_workbook(tmp_path / "input.xlsx")])
    assert "target_price" not in result["missing_fields"]
    assert result["component_costing_ready"] is True


def test_candidate_provenance_is_retained():
    result = merge_customer_input_candidates([{
        "field": "customer",
        "value": "A",
        "source_file": "input.csv",
        "source_sheet": "CSV",
        "source_cell": "R1C2",
        "raw_label": "Customer",
        "raw_value": "A",
        "confidence": "high",
        "priority": 3,
        "is_formula": False,
        "formula": None,
        "source_type": "commercial_csv",
    }])
    selected = result["fields"]["customer"]["selected_candidate"]
    assert selected["source_file"] == "input.csv"
    assert selected["source_cell"] == "R1C2"


def test_legacy_currency_mirrors_quotation_currency_only():
    normalized = normalize_customer_input({
        "quotation_currency": "INR",
        "target_price_currency": "EUR",
    })["customer_input"]
    assert normalized["currency"] == "INR"
    assert normalized["target_price_currency"] == "EUR"


def _xlsx_bytes():
    workbook = openpyxl.Workbook()
    stream = BytesIO()
    try:
        sheet = workbook.active
        sheet.append(["Product", "Rod Choke"])
        sheet.append(["Annual quantity", 100000])
        sheet.append(["Delivery country", "India"])
        sheet.append(["Quotation currency", "INR"])
        workbook.save(stream)
        return stream.getvalue()
    finally:
        workbook.close()
        stream.close()


def test_excel_reader_closes_formula_value_and_vba_archives(monkeypatch, tmp_path):
    from services import customer_input_extraction as extraction

    events = []

    class FakeArchive:
        def __init__(self, name):
            self.name = name

        def close(self):
            events.append(f"{self.name}.vba.close")

    class FakeWorkbook:
        def __init__(self, name):
            self.name = name
            self.vba_archive = FakeArchive(name)

        def close(self):
            events.append(f"{self.name}.workbook.close")

    workbooks = [FakeWorkbook("formulas"), FakeWorkbook("values")]
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: workbooks.pop(0))
    monkeypatch.setattr(
        extraction,
        "_extract_excel_fields_from_workbooks",
        lambda *args: {"status": "ok"},
    )

    assert extraction.extract_excel_fields(tmp_path / "input.xlsm") == {"status": "ok"}
    assert events == [
        "values.workbook.close",
        "values.vba.close",
        "formulas.workbook.close",
        "formulas.vba.close",
    ]


def test_green_commercial_cells_are_mapped_with_full_provenance(tmp_path):
    workbook = openpyxl.Workbook()
    path = tmp_path / "green-inputs.xlsx"
    try:
        sheet = workbook.active
        sheet.title = "Commercial"
        green = openpyxl.styles.PatternFill(
            fill_type="solid", fgColor="00B050"
        )
        sheet["A1"] = "Payment terms"
        sheet["B1"] = 60
        sheet["B1"].fill = green
        sheet["B1"].number_format = "0"
        sheet["A2"] = "Incoterm"
        sheet["B2"] = "FCA"
        sheet["B2"].fill = green
        sheet["A3"] = "Unmapped commercial note"
        sheet["B3"] = "Special review"
        sheet["B3"].fill = green
        workbook.save(path)
    finally:
        workbook.close()

    report = extract_excel_fields(path)
    mapped = {
        item["normalized_destination_field"]: item
        for item in report["green_commercial_inputs"]
    }
    assert mapped["customer_payment_days"]["value"] == 60
    assert mapped["customer_payment_days"]["cell_address"] == "B1"
    assert mapped["customer_payment_days"]["displayed_label"] == "Payment terms"
    assert mapped["customer_payment_days"]["number_format"] == "0"
    assert mapped["customer_payment_days"]["fill_color"]
    assert mapped["customer_incoterm"]["normalized_value"] == "FCA"
    assert report["unmapped_green_inputs"][0]["value"] == "Special review"
    assert report["unmapped_green_inputs"][0][
        "normalized_destination_field"
    ] is None


def test_green_inputs_are_exposed_at_package_level(tmp_path):
    workbook = openpyxl.Workbook()
    path = tmp_path / "commercial.xlsx"
    try:
        sheet = workbook.active
        sheet["A1"] = "Target price"
        sheet["B1"] = 12.5
        sheet["B1"].fill = openpyxl.styles.PatternFill(
            fill_type="solid", fgColor="92D050"
        )
        workbook.save(path)
    finally:
        workbook.close()
    result = extract_customer_input_package({}, [path])
    assert result["green_commercial_inputs"][0][
        "normalized_destination_field"
    ] == "target_price"


def _client_with_temp_inputs(monkeypatch, tmp_path):
    from app.main import app
    from app.routers import choke_costing_ui_router as router_module
    from services import choke_sequential_agent_workflow as workflow_module

    monkeypatch.setattr(router_module, "CUSTOMER_INPUT_DIR", tmp_path)
    monkeypatch.setattr(workflow_module, "CUSTOMER_INPUT_DIR", tmp_path)
    monkeypatch.setattr(router_module, "append_workflow_event", lambda *args, **kwargs: None)
    monkeypatch.setattr(router_module, "is_azure_blob_configured", lambda: False)
    return TestClient(app)


def test_api_multipart_drawing_pdf_only(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    response = client.post("/api/choke-costing/customer-inputs/create", files={
        "drawing_pdf": ("drawing.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"),
    })
    assert response.status_code == 200
    assert len(response.json()["attachment_manifest"]) == 1


def test_api_multipart_pdf_plus_xlsm(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    response = client.post("/api/choke-costing/customer-inputs/create", files=[
        ("drawing_pdf", ("drawing.pdf", b"%PDF-1.4\n%%EOF", "application/pdf")),
        ("attachments", ("rfq.xlsm", _xlsx_bytes(), "application/vnd.ms-excel.sheet.macroEnabled.12")),
    ])
    assert response.status_code == 200
    assert response.json()["component_costing_ready"] is True
    saved = response.json()["customer_input"]
    resolution = client.get(
        f"/api/choke-workflow/customer-input-resolution/{saved['project_code']}/{saved['product_id']}"
    )
    assert resolution.status_code == 200
    assert resolution.json()["component_costing_ready"] is True


def test_api_repeated_attachment_entries_are_all_saved(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    response = client.post("/api/choke-costing/customer-inputs/create", files=[
        ("drawing_pdf", ("drawing.pdf", b"%PDF-1.4\n%%EOF", "application/pdf")),
        ("attachments", ("rfq.xlsx", _xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ("attachments", ("volume.csv", b"Annual quantity,100000", "text/csv")),
    ])
    assert response.status_code == 200
    assert len(response.json()["attachment_manifest"]) == 3


def test_api_manifest_contains_every_file_and_metadata(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    response = client.post("/api/choke-costing/customer-inputs/create", files=[
        ("drawing_pdf", ("same.pdf", b"%PDF-one", "application/pdf")),
        ("attachments", ("same.pdf", b"%PDF-two", "application/pdf")),
    ])
    manifest = response.json()["attachment_manifest"]
    assert len(manifest) == 2
    assert manifest[0]["stored_filename"] != manifest[1]["stored_filename"]
    assert all(item["checksum_sha256"] and item["file_size"] for item in manifest)


def test_api_unsupported_attachment_returns_clear_4xx(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    response = client.post("/api/choke-costing/customer-inputs/create", files=[
        ("drawing_pdf", ("drawing.pdf", b"%PDF-1.4\n%%EOF", "application/pdf")),
        ("attachments", ("malware.exe", b"MZ", "application/octet-stream")),
    ])
    assert response.status_code == 400
    assert "Unsupported attachment type" in response.json()["detail"]


def test_api_reuses_identical_files_for_same_saved_project(monkeypatch, tmp_path):
    client = _client_with_temp_inputs(monkeypatch, tmp_path)
    files = [
        ("drawing_pdf", ("drawing.pdf", b"%PDF-same", "application/pdf")),
        ("attachments", ("rfq.csv", b"Annual quantity,100000", "text/csv")),
    ]
    first = client.post(
        "/api/choke-costing/customer-inputs/create",
        data={"project_code": "REUSE-TEST", "product_id": "PART-1"},
        files=files,
    )
    second = client.post(
        "/api/choke-costing/customer-inputs/create",
        data={"project_code": "REUSE-TEST", "product_id": "PART-1"},
        files=files,
    )
    assert first.status_code == second.status_code == 200
    assert all(item["reused_existing_file"] is True for item in second.json()["attachment_manifest"])
    assert len(list((tmp_path / "uploads" / "REUSE-TEST").iterdir())) == 2


def test_choke_family_pluralization_and_case_are_equivalent():
    candidates = []
    for index, value in enumerate(["Choke", "Chokes", "choke", "chokes"]):
        candidates.append({
            "field": "product_family", "value": value, "source_file": f"source-{index}",
            "source_sheet": None, "source_cell": None, "raw_label": "Product family",
            "raw_value": value, "confidence": "high", "priority": index + 1,
            "is_formula": False, "formula": None, "source_type": "structured_request_json",
        })
    result = merge_customer_input_candidates(candidates)
    assert result["fields"]["product_family"]["value"] == "Chokes"
    assert result["conflicts"] == []
    assert normalize_product_family("choke") == "Chokes"


def test_drawing_paths_and_collision_safe_names_are_equivalent():
    excel = "0300440157_INDUCTOR_20.04.24.pdf"
    stored = "https://host/files/project/0300440157_INDUCTOR_20_04_24__2.pdf"
    assert normalize_drawing_reference(excel) == normalize_drawing_reference(stored)


def test_genuinely_different_drawing_references_remain_blocking():
    candidates = []
    for index, value in enumerate(["0300440157_INDUCTOR_20.04.24.pdf", "OTHER_PART_DRAWING.pdf"]):
        candidates.append({
            "field": "drawing_reference", "value": value, "source_file": f"source-{index}",
            "source_sheet": None, "source_cell": None, "raw_label": "Drawing reference",
            "raw_value": value, "confidence": "high", "priority": index + 1,
            "is_formula": False, "formula": None, "source_type": "structured_request_json",
        })

    result = merge_customer_input_candidates(candidates)

    assert result["conflicts"][0]["field"] == "drawing_reference"
    assert result["conflicts"][0]["blocking"] is True


def test_canonical_commercial_payload_passes_backend_schema():
    from app.routers.choke_workflow_router import UpdateCommercialFieldsRequest

    request = UpdateCommercialFieldsRequest.model_validate({
        "project_code": "24018-CHO-00",
        "product_id": 300440157,
        "product": "Rod Choke",
        "product_name": "Rod Choke",
        "product_family": "Chokes",
        "part_number": 300440157,
        "annual_quantity": 360000,
        "customer_delivery_zone": "India",
        "quotation_currency": "INR",
        "drawing_reference": "0300440157_INDUCTOR_20.04.24.pdf",
    })
    assert request.product_id == "300440157"
    assert request.part_number == "300440157"
    assert request.annual_quantity == 360000
