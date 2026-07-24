"""Generic, provenance-preserving customer-input extraction.

The service never executes workbook macros and never treats hidden financial
content or stale external links as authoritative commercial input.
"""

from __future__ import annotations

import csv
import json
import mimetypes
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import unquote, urlsplit

from services.currency_service import normalize_currency_code
from services.manufacturing_strategy import load_product_matrix, normalize_delivery_zone
from services.project_data_paths import data_reference_candidates, portable_data_reference


SUPPORTED_CUSTOMER_INPUT_EXTENSIONS = {".pdf", ".xlsx", ".xlsm", ".csv"}
REQUIRED_COMPONENT_COSTING_FIELDS = {
    "product_name": "Complete product identification",
    "annual_quantity": "Annual quantity",
    "delivery_zone": "Delivery destination/zone",
    "quotation_currency": "Quotation currency",
}
BLOCKING_CONFLICT_FIELDS = set(REQUIRED_COMPONENT_COSTING_FIELDS) | {
    "product_family",
    "drawing_reference",
}

FIELD_ALIASES = {
    "project_code": ["quote id", "rfq number", "rfq id", "project code", "project number"],
    "customer": ["customer", "customer name", "client"],
    "final_customer": ["final customer", "end customer"],
    "project_name": ["project name", "program name"],
    "product_name": ["product", "product name", "product type"],
    "product_family": ["product line", "product family", "family"],
    "product_subtype": ["product subtype", "choke subtype"],
    "part_number": ["part number", "part no", "part no.", "customer part number", "customer p/n"],
    "annual_quantity": ["annual quantity", "annual volume", "yearly volume", "qmax", "maximum annual quantity"],
    "sop_date": ["sop", "sop date", "start of production"],
    "delivery_country": ["delivery country", "country", "delivery area"],
    "delivery_city": ["delivery city", "city/country", "ship to city"],
    "delivery_zone": ["delivery zone", "destination zone", "customer delivery zone"],
    "production_plant": ["production plant", "manufacturing plant", "plant to produce"],
    "target_price": ["target price", "customer target price", "cust target price"],
    "target_price_currency": ["target price currency"],
    "quotation_currency": ["quotation currency", "quote currency", "selling currency"],
    "purchasing_currency": ["purchasing currency", "purchase currency"],
    "drawing_reference": ["drawing reference", "drawing filename", "drawing file"],
    "customer_payment_days": [
        "payment terms", "customer payment terms", "payment conditions",
    ],
    "customer_incoterm": ["incoterm", "delivery conditions", "delivery condition"],
    "customer_delivery_frequency_days": [
        "delivery frequency", "delivery frequency days",
    ],
    "platform": ["platform", "delivery platform"],
    "customer_productivity_percentage": [
        "customer productivity", "productivity percentage", "productivity",
    ],
    "productivity_basis": ["productivity basis"],
    "capex_tooling_treatment": [
        "capex treatment", "tooling treatment", "capex tooling treatment",
    ],
    "profitability_target": [
        "profitability target", "target roce", "roce target", "npv target",
    ],
    "indexation": ["indexation", "material indexation", "price indexation"],
}

STRUCTURED_ALIASES = {
    "project_code": ["project_code", "project_number", "systematic_rfq_id", "rfq_number"],
    "customer": ["customer", "customer_name"],
    "final_customer": ["final_customer", "end_customer"],
    "project_name": ["project_name"],
    "product_name": ["product", "product_name"],
    "product_family": ["product_family", "product_line", "product_line_name"],
    "product_subtype": ["product_subtype", "choke_subtype"],
    "part_number": ["part_number", "customer_part_number", "product_reference"],
    "annual_quantity": ["annual_quantity", "annual_volume", "qmax"],
    "sop_date": ["sop_date", "sop", "sop_year"],
    "delivery_country": ["delivery_country", "country"],
    "delivery_city": ["delivery_city", "city"],
    "delivery_zone": ["customer_delivery_zone", "delivery_zone", "destination_zone"],
    "production_plant": ["production_plant"],
    "target_price": ["target_price", "target_price_value"],
    "target_price_currency": ["target_price_currency"],
    "quotation_currency": ["quotation_currency", "selling_currency", "currency"],
    "purchasing_currency": ["purchasing_currency", "purchasing_currency_code"],
    "drawing_reference": ["drawing_original_filename", "drawing_reference", "drawing_file", "drawing"],
}

COUNTRY_ZONE_MAP = {
    "india": "India",
    "china": "China South Pacific",
    "pr china": "China South Pacific",
    "people s republic of china": "China South Pacific",
    "tunisia": "Africa",
    "france": "Europe",
}


def _text(value: Any) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_product_family(value: Any) -> Optional[str]:
    key = _text(value)
    if key in {"choke", "chokes"}:
        return "Chokes"
    return str(value).strip() if _has_value(value) else None


def normalize_drawing_reference(value: Any) -> str:
    text = unquote(str(value or "").strip())
    path = urlsplit(text).path.replace("\\", "/")
    basename = Path(path).name.strip().lower()
    stem = Path(basename).stem
    stem = re.sub(r"__\d+$", "", stem)
    suffix = Path(basename).suffix.lower()
    return re.sub(r"[^a-z0-9]+", "", stem) + re.sub(r"[^a-z0-9]+", "", suffix)


def _has_value(value: Any) -> bool:
    return value not in (None, "", [], {}) and str(value).strip() not in {".", "-", "n/a", "N/A"}


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _number(value: Any) -> Optional[float]:
    if isinstance(value, bool) or value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    text = re.sub(r"(?i)(pcs?|pieces?|units?|peryear|/year|pa)$", "", text)
    try:
        return float(text)
    except ValueError:
        return None


def _field_from_label(value: Any) -> Optional[str]:
    label = _text(value)
    if not label:
        return None
    # More specific aliases must win over their shorter parent labels.
    ordered = sorted(
        ((field, alias) for field, aliases in FIELD_ALIASES.items() for alias in aliases),
        key=lambda item: len(item[1]),
        reverse=True,
    )
    for field, alias in ordered:
        alias_text = _text(alias)
        if label == alias_text or label.startswith(f"{alias_text} "):
            return field
    return None


def _candidate(
    field: str,
    value: Any,
    *,
    source_file: str,
    source_sheet: Optional[str] = None,
    source_cell: Optional[str] = None,
    raw_label: Optional[str] = None,
    raw_value: Any = None,
    confidence: str = "high",
    priority: int = 3,
    is_formula: bool = False,
    formula: Optional[str] = None,
    source_type: str = "visible_commercial_workbook",
    **extra: Any,
) -> Dict[str, Any]:
    return {
        "field": field,
        "value": _json_value(value),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_cell": source_cell,
        "raw_label": raw_label,
        "raw_value": _json_value(value if raw_value is None else raw_value),
        "confidence": confidence,
        "priority": priority,
        "is_formula": bool(is_formula),
        "formula": formula,
        "source_type": source_type,
        **extra,
    }


def _fill_color(cell: Any) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    color = getattr(fill, "fgColor", None)
    if fill is None or getattr(fill, "fill_type", None) not in {"solid", "pattern"}:
        return None
    color_type = getattr(color, "type", None)
    value = getattr(color, "rgb", None)
    if color_type == "indexed":
        value = f"indexed:{getattr(color, 'indexed', None)}"
    elif color_type == "theme":
        value = f"theme:{getattr(color, 'theme', None)}:{getattr(color, 'tint', 0)}"
    return str(value) if value not in (None, "") else None


def _is_green_fill(cell: Any) -> bool:
    color = _fill_color(cell)
    if not color:
        return False
    if color.startswith("indexed:"):
        return color in {"indexed:17", "indexed:42", "indexed:43"}
    if color.startswith("theme:"):
        return False
    rgb = color[-6:]
    try:
        red, green, blue = (
            int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
        )
    except (ValueError, IndexError):
        return False
    return green >= 96 and green > red * 1.15 and green > blue * 1.10


def _green_cell_label(sheet: Any, cell: Any) -> tuple[Optional[str], Optional[str]]:
    candidates = []
    for row_offset, column_offset in ((0, -1), (0, -2), (-1, 0), (-2, 0)):
        row = cell.row + row_offset
        column = cell.column + column_offset
        if row < 1 or column < 1:
            continue
        label_cell = sheet.cell(row, column)
        if _has_value(label_cell.value):
            candidates.append((str(label_cell.value), label_cell.coordinate))
    return candidates[0] if candidates else (None, None)


def _normalize_green_value(field: str, value: Any) -> Any:
    if field in {
        "annual_quantity", "target_price", "customer_delivery_frequency_days",
        "customer_productivity_percentage", "profitability_target",
    }:
        return _number(value)
    if field == "customer_payment_days":
        numeric = _number(value)
        if numeric is not None:
            return numeric
        match = re.search(r"\d+(?:[.,]\d+)?", str(value or ""))
        return _number(match.group(0)) if match else None
    if field in {
        "quotation_currency", "target_price_currency", "purchasing_currency",
    }:
        return normalize_currency_code(value)
    if field == "platform":
        key = _text(value)
        if key in {"yes", "y", "true", "1", "oui"}:
            return True
        if key in {"no", "n", "false", "0", "non"}:
            return False
    return _json_value(value)


def _resolve_stored_path(value: Any) -> Optional[Path]:
    if not value:
        return None
    path = Path(str(value))
    for candidate in data_reference_candidates(path):
        candidate = candidate.resolve()
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def discover_customer_input_files(source: Any) -> List[Dict[str, Any]]:
    """Return supported files from a directory, path list, or attachment manifest."""
    entries: List[Dict[str, Any]] = []
    if source is None:
        return entries
    if isinstance(source, (str, Path)):
        path = Path(source)
        raw_items: List[Any] = list(path.iterdir()) if path.is_dir() else [path]
    elif isinstance(source, dict):
        raw_items = source.get("attachments") or source.get("attachment_manifest") or [source]
    else:
        raw_items = list(source)

    seen = set()
    for item in raw_items:
        manifest = item if isinstance(item, dict) else {}
        path = (
            _resolve_stored_path(manifest.get("stored_path") or manifest.get("path"))
            if manifest else Path(item).resolve()
        )
        if not path or not path.exists() or not path.is_file():
            continue
        extension = path.suffix.lower()
        if extension not in SUPPORTED_CUSTOMER_INPUT_EXTENSIONS:
            continue
        identity = str(path.resolve()).lower()
        if identity in seen:
            continue
        seen.add(identity)
        entries.append({
            **manifest,
            "path": path,
            "stored_path": manifest.get("stored_path") or portable_data_reference(path),
            "original_filename": manifest.get("original_filename") or path.name,
            "stored_filename": manifest.get("stored_filename") or path.name,
            "mime_type": manifest.get("mime_type") or mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            "file_size": manifest.get("file_size") or path.stat().st_size,
            "extension": extension,
        })
    return sorted(entries, key=lambda item: (item["extension"], item["original_filename"].lower()))


def _cell_value(formula_sheet: Any, value_sheet: Any, cell: Any) -> tuple[Any, bool, Optional[str]]:
    is_formula = cell.data_type == "f"
    formula = str(cell.value) if is_formula else None
    value = value_sheet[cell.coordinate].value if is_formula else cell.value
    return _json_value(value), is_formula, formula


def _nearby_value(
    formula_sheet: Any,
    value_sheet: Any,
    label_cell: Any,
    field: str,
) -> Optional[Dict[str, Any]]:
    table_header = bool(getattr(label_cell.fill, "fill_type", None) or label_cell.font.bold)
    positions = (
        [(offset, 0) for offset in range(1, 5)]
        if table_header
        else [(0, offset) for offset in range(1, 5)] + [(offset, 0) for offset in range(1, 5)]
    )
    for row_offset, column_offset in positions:
        row = label_cell.row + row_offset
        column = label_cell.column + column_offset
        if row > formula_sheet.max_row or column > formula_sheet.max_column:
            continue
        cell = formula_sheet.cell(row, column)
        value, is_formula, formula = _cell_value(formula_sheet, value_sheet, cell)
        if not _has_value(value) or _field_from_label(value):
            continue
        same_style = cell.style_id == label_cell.style_id
        label_fill = getattr(label_cell.fill, "fill_type", None)
        same_nondefault_fill = bool(
            label_fill
            and label_fill == getattr(cell.fill, "fill_type", None)
            and label_cell.fill.fgColor.type == cell.fill.fgColor.type
            and label_cell.fill.fgColor.rgb == cell.fill.fgColor.rgb
        )
        if isinstance(value, str) and ((same_style and label_cell.style_id != 0) or same_nondefault_fill):
            continue
        if field in {"target_price", "annual_quantity"} and _number(value) is None:
            continue
        if field in {"delivery_country", "delivery_city", "delivery_zone"} and isinstance(value, (int, float)):
            continue
        return {"cell": cell.coordinate, "value": value, "is_formula": is_formula, "formula": formula}
    return None


def _yearly_volume_candidates(formula_sheet: Any, value_sheet: Any, source_file: str) -> List[Dict[str, Any]]:
    results = []
    for row_index in range(1, formula_sheet.max_row + 1):
        years = []
        for column in range(1, formula_sheet.max_column + 1):
            value = formula_sheet.cell(row_index, column).value
            year = int(value) if isinstance(value, (int, float)) and 2000 <= int(value) <= 2200 else None
            if year:
                years.append((column, year))
        if len(years) < 2:
            continue
        context = " ".join(
            str(formula_sheet.cell(row, column).value or "")
            for row in range(max(1, row_index - 3), row_index + 1)
            for column in range(1, formula_sheet.max_column + 1)
        )
        if not re.search(r"(?i)\b(volume|annual quantity|annual volume|qmax)\b", context):
            continue
        for value_row in range(row_index + 1, min(formula_sheet.max_row, row_index + 3) + 1):
            series = {}
            formula_flags = []
            for column, year in years:
                cell = formula_sheet.cell(value_row, column)
                value, is_formula, formula = _cell_value(formula_sheet, value_sheet, cell)
                number = _number(value)
                if number is not None and number >= 0:
                    series[str(year)] = int(number) if number.is_integer() else number
                    formula_flags.append((is_formula, formula))
            positive = [float(value) for value in series.values() if float(value) > 0]
            if not positive:
                continue
            qmax = max(positive)
            start = formula_sheet.cell(row_index, years[0][0]).coordinate
            end = formula_sheet.cell(value_row, years[-1][0]).coordinate
            results.append(_candidate(
                "annual_quantity",
                int(qmax) if qmax.is_integer() else qmax,
                source_file=source_file,
                source_sheet=formula_sheet.title,
                source_cell=f"{start}:{end}",
                raw_label="Yearly volume series",
                raw_value=series,
                confidence="high",
                priority=3,
                is_formula=any(item[0] for item in formula_flags),
                formula="; ".join(item[1] for item in formula_flags if item[1]) or None,
                quantity_by_year=series,
                unit="pcs/year",
                qmax=int(qmax) if qmax.is_integer() else qmax,
                derivation_method="maximum_positive_annual_volume_not_sum",
            ))
            break
    return results


def _close_openpyxl_workbook(workbook: Any) -> None:
    if workbook is None:
        return
    archive = getattr(workbook, "_archive", None)
    vba_archive = getattr(workbook, "vba_archive", None)
    try:
        workbook.close()
    finally:
        if vba_archive is not None and vba_archive is not archive:
            vba_archive.close()


def extract_excel_fields(path: str | Path) -> Dict[str, Any]:
    import openpyxl

    workbook_path = Path(path).resolve()
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    formulas = None
    values = None
    try:
        formulas = openpyxl.load_workbook(
            workbook_path, data_only=False, keep_vba=keep_vba, keep_links=True
        )
        values = openpyxl.load_workbook(
            workbook_path, data_only=True, keep_vba=keep_vba, keep_links=True
        )
        return _extract_excel_fields_from_workbooks(workbook_path, formulas, values)
    finally:
        _close_openpyxl_workbook(values)
        _close_openpyxl_workbook(formulas)


def _extract_excel_fields_from_workbooks(
    workbook_path: Path,
    formulas: Any,
    values: Any,
) -> Dict[str, Any]:
    candidates: List[Dict[str, Any]] = []
    warnings: List[str] = []
    sheet_inventory = []
    green_commercial_inputs: List[Dict[str, Any]] = []
    unmapped_green_inputs: List[Dict[str, Any]] = []

    external_links = len(getattr(formulas, "_external_links", []) or [])
    if external_links:
        warnings.append(f"Ignored {external_links} unavailable/stale external workbook link(s) in {workbook_path.name}.")

    ignored_names = []
    for name in formulas.defined_names.values():
        broken = "#REF!" in str(name.attr_text or "")
        destinations = []
        if not broken:
            try:
                destinations = list(name.destinations)
            except Exception:
                destinations = []
        if broken or any(sheet not in formulas.sheetnames for sheet, _ in destinations):
            ignored_names.append(name.name)
            continue
        field = _field_from_label(name.name)
        if not field:
            continue
        for sheet_name, coordinate in destinations:
            if formulas[sheet_name].sheet_state != "visible":
                continue
            coordinate = coordinate.replace("$", "")
            if ":" in coordinate:
                continue
            formula_cell = formulas[sheet_name][coordinate]
            value, is_formula, formula = _cell_value(
                formulas[sheet_name], values[sheet_name], formula_cell
            )
            if field == "product_family":
                value = normalize_product_family(value)
            elif field in {"quotation_currency", "target_price_currency", "purchasing_currency"}:
                value = normalize_currency_code(value)
            elif field in {"annual_quantity", "target_price"}:
                value = _number(value)
            if _has_value(value):
                candidates.append(_candidate(
                    field,
                    value,
                    source_file=workbook_path.name,
                    source_sheet=sheet_name,
                    source_cell=coordinate,
                    raw_label=name.name,
                    raw_value=value,
                    confidence="high",
                    priority=3,
                    is_formula=is_formula,
                    formula=formula,
                    source_type="visible_commercial_named_range",
                ))
    if ignored_names:
        warnings.append(
            f"Ignored {len(ignored_names)} broken or unavailable named range(s) in "
            f"{workbook_path.name}: {', '.join(ignored_names)}."
        )

    for sheet in formulas.worksheets:
        sheet_inventory.append({
            "sheet": sheet.title,
            "state": sheet.sheet_state,
            "used_range": sheet.calculate_dimension(),
            "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
            "tables": list(sheet.tables.keys()),
        })
        if sheet.sheet_state != "visible":
            warnings.append(f"Inspected but did not use hidden sheet '{sheet.title}' as commercial input.")
            continue
        value_sheet = values[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                if not _is_green_fill(cell):
                    continue
                value, is_formula, formula = _cell_value(
                    sheet, value_sheet, cell
                )
                if not _has_value(value):
                    continue
                displayed_label, label_cell = _green_cell_label(sheet, cell)
                field = _field_from_label(displayed_label)
                record = {
                    "workbook_filename": workbook_path.name,
                    "sheet": sheet.title,
                    "cell_address": cell.coordinate,
                    "displayed_label": displayed_label,
                    "label_cell": label_cell,
                    "value": _json_value(value),
                    "number_format": cell.number_format,
                    "fill_color": _fill_color(cell),
                    "is_formula": is_formula,
                    "formula": formula,
                    "extraction_confidence": "high" if field else "unmapped",
                    "normalized_destination_field": field,
                }
                if not field:
                    unmapped_green_inputs.append(record)
                    continue
                normalized_value = _normalize_green_value(field, value)
                if not _has_value(normalized_value):
                    record["extraction_confidence"] = "low"
                    record["mapping_warning"] = (
                        "Recognized label, but the value could not be normalized."
                    )
                    unmapped_green_inputs.append(record)
                    continue
                record["normalized_value"] = normalized_value
                green_commercial_inputs.append(record)
                candidates.append(_candidate(
                    field,
                    normalized_value,
                    source_file=workbook_path.name,
                    source_sheet=sheet.title,
                    source_cell=cell.coordinate,
                    raw_label=displayed_label,
                    raw_value=value,
                    confidence="high",
                    priority=3,
                    is_formula=is_formula,
                    formula=formula,
                    source_type="green_commercial_workbook_input",
                    number_format=cell.number_format,
                    fill_color=_fill_color(cell),
                    label_cell=label_cell,
                ))
        for row in sheet.iter_rows():
            for cell in row:
                field = _field_from_label(cell.value)
                if not field:
                    continue
                nearby = _nearby_value(sheet, value_sheet, cell, field)
                if not nearby:
                    continue
                value = nearby["value"]
                if field == "product_family":
                    value = normalize_product_family(value)
                elif field in {"quotation_currency", "target_price_currency", "purchasing_currency"}:
                    value = normalize_currency_code(value)
                elif field in {"target_price", "annual_quantity"}:
                    value = _number(value)
                if not _has_value(value):
                    continue
                candidates.append(_candidate(
                    field,
                    value,
                    source_file=workbook_path.name,
                    source_sheet=sheet.title,
                    source_cell=nearby["cell"],
                    raw_label=str(cell.value),
                    raw_value=nearby["value"],
                    confidence="high",
                    priority=3,
                    is_formula=nearby["is_formula"],
                    formula=nearby["formula"],
                ))
                
                # Keep scanning: repeated fields are valuable conflict evidence.
        candidates.extend(_yearly_volume_candidates(sheet, value_sheet, workbook_path.name))
        for row in sheet.iter_rows():
            for cell in row:
                value, is_formula, formula = _cell_value(sheet, value_sheet, cell)
                if isinstance(value, str) and Path(value.strip()).suffix.lower() == ".pdf":
                    candidates.append(_candidate(
                        "drawing_reference",
                        Path(value.strip()).name,
                        source_file=workbook_path.name,
                        source_sheet=sheet.title,
                        source_cell=cell.coordinate,
                        raw_label="PDF document reference",
                        raw_value=value,
                        confidence="high",
                        priority=3,
                        is_formula=is_formula,
                        formula=formula,
                    ))

    result = {
        "source_file": workbook_path.name,
        "source_path": str(workbook_path),
        "candidates": candidates,
        "warnings": list(dict.fromkeys(warnings)),
        "sheets": sheet_inventory,
        "green_commercial_inputs": green_commercial_inputs,
        "unmapped_green_inputs": unmapped_green_inputs,
        "macros_executed": False,
    }
    return result


def extract_csv_fields(path: str | Path) -> Dict[str, Any]:
    csv_path = Path(path).resolve()
    with csv_path.open("r", encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    candidates: List[Dict[str, Any]] = []
    for row_index, row in enumerate(rows, start=1):
        for column_index, label in enumerate(row, start=1):
            field = _field_from_label(label)
            if not field:
                continue
            value = next((item for item in row[column_index:] if _has_value(item) and not _field_from_label(item)), None)
            if value is None and row_index < len(rows):
                below = rows[row_index]
                value = below[column_index - 1] if column_index - 1 < len(below) else None
            if not _has_value(value):
                continue
            if field == "product_family":
                value = normalize_product_family(value)
            elif field in {"quotation_currency", "target_price_currency", "purchasing_currency"}:
                value = normalize_currency_code(value)
            elif field in {"annual_quantity", "target_price"}:
                value = _number(value)
            if not _has_value(value):
                continue
            candidates.append(_candidate(
                field,
                value,
                source_file=csv_path.name,
                source_sheet="CSV",
                source_cell=f"R{row_index}C{column_index}",
                raw_label=label,
                raw_value=value,
                confidence="high",
                priority=3,
                source_type="commercial_csv",
            ))
    return {"source_file": csv_path.name, "source_path": str(csv_path), "candidates": candidates, "warnings": []}


def extract_pdf_metadata(path: str | Path) -> Dict[str, Any]:
    pdf_path = Path(path).resolve()
    return {
        "source_file": pdf_path.name,
        "source_path": str(pdf_path),
        "candidates": [_candidate(
            "drawing_reference",
            pdf_path.name,
            source_file=pdf_path.name,
            source_cell="filename",
            raw_label="PDF filename",
            raw_value=pdf_path.name,
            confidence="high",
            priority=4,
            source_type="drawing_specification",
        )],
        "warnings": [],
    }


def extract_structured_request_fields(payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    payload = payload or {}
    explicit_fields = set(payload.get("explicit_user_confirmed_fields") or payload.get("_explicit_user_fields") or [])
    candidates = []
    for field, aliases in STRUCTURED_ALIASES.items():
        source_key = next((key for key in aliases if _has_value(payload.get(key))), None)
        if not source_key:
            continue
        value = payload[source_key]
        if field == "product_family":
            value = normalize_product_family(value)
        elif field in {"quotation_currency", "target_price_currency", "purchasing_currency"}:
            value = normalize_currency_code(value)
        elif field in {"annual_quantity", "target_price"}:
            value = _number(value)
        if not _has_value(value):
            continue
        explicit = source_key in explicit_fields or field in explicit_fields
        candidates.append(_candidate(
            field,
            value,
            source_file="request_payload",
            source_cell=source_key,
            raw_label=source_key,
            raw_value=payload[source_key],
            confidence="confirmed" if explicit else "high",
            priority=1 if explicit else 2,
            source_type="explicit_user_confirmed" if explicit else "structured_request_json",
        ))
    return {"source_file": "request_payload", "candidates": candidates, "warnings": []}


def _zone_vocabulary() -> List[str]:
    zones = []
    for row in load_product_matrix():
        zones.extend(row.get("zones", {}).keys())
    return list(dict.fromkeys(str(item).strip() for item in zones if str(item).strip()))


def normalize_delivery_location(
    country: Any = None,
    city: Any = None,
    zone: Any = None,
) -> Dict[str, Any]:
    vocabulary = _zone_vocabulary()
    if _has_value(zone):
        normalized = normalize_delivery_zone(zone)
        match = next((item for item in vocabulary if _text(item) == _text(normalized)), normalized)
        return {
            "raw_country": country,
            "raw_city": city,
            "raw_zone": zone,
            "normalized_zone": match,
            "mapping_source": "explicit_zone_existing_vocabulary",
            "mapping_rule": f"explicit:{zone}",
            "confidence": "high",
        }
    country_key = _text(country)
    mapped = COUNTRY_ZONE_MAP.get(country_key)
    if mapped:
        match = next((item for item in vocabulary if _text(item) == _text(mapped)), mapped)
        return {
            "raw_country": country,
            "raw_city": city,
            "raw_zone": None,
            "normalized_zone": match,
            "mapping_source": "country_to_avocarbon_zone",
            "mapping_rule": f"{country}->{match}",
            "confidence": "high",
        }
    return {
        "raw_country": country,
        "raw_city": city,
        "raw_zone": zone,
        "normalized_zone": None,
        "mapping_source": None,
        "mapping_rule": None,
        "confidence": "low",
    }


def _comparable_value(field: str, value: Any) -> Any:
    if field in {"annual_quantity", "target_price"}:
        return _number(value)
    if field.endswith("currency"):
        return normalize_currency_code(value)
    if field == "product_family":
        return _text(normalize_product_family(value))
    if field == "drawing_reference":
        return normalize_drawing_reference(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return _text(value)


def merge_customer_input_candidates(
    candidates: Iterable[Dict[str, Any]],
    warnings: Optional[Iterable[str]] = None,
) -> Dict[str, Any]:
    all_candidates = [dict(item) for item in candidates if _has_value(item.get("value"))]
    for item in all_candidates:
        if item.get("field") == "product_family":
            item["value"] = normalize_product_family(item.get("value"))
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for item in all_candidates:
        grouped.setdefault(item["field"], []).append(item)

    # A country/city becomes a zone candidate only after preserving its raw provenance.
    country = min(grouped.get("delivery_country", []), key=lambda item: item["priority"], default={}).get("value")
    city = min(grouped.get("delivery_city", []), key=lambda item: item["priority"], default={}).get("value")
    explicit_zone = min(grouped.get("delivery_zone", []), key=lambda item: item["priority"], default={}).get("value")
    location = normalize_delivery_location(country, city, explicit_zone)
    if location.get("normalized_zone"):
        source = min(
            grouped.get("delivery_zone", []) or grouped.get("delivery_country", []),
            key=lambda item: item["priority"],
        )
        grouped.setdefault("delivery_zone", []).append(_candidate(
            "delivery_zone",
            location["normalized_zone"],
            source_file=source.get("source_file") or "delivery_normalization",
            source_sheet=source.get("source_sheet"),
            source_cell=source.get("source_cell"),
            raw_label=source.get("raw_label"),
            raw_value=source.get("value"),
            confidence=location["confidence"],
            priority=source["priority"],
            source_type="delivery_zone_normalization",
            delivery_location=location,
        ))

    fields = {}
    resolved_fields = []
    conflicts = []
    for field, items in grouped.items():
        items.sort(key=lambda item: (item["priority"], str(item.get("source_file")), str(item.get("source_cell"))))
        selected = items[0]
        unique = {}
        for item in items:
            unique.setdefault(str(_comparable_value(field, item["value"])), item)
        conflict_candidates = list(unique.values())
        explicit_confirmation = selected.get("source_type") == "explicit_user_confirmed"
        if len(conflict_candidates) > 1 and not explicit_confirmation:
            conflicts.append({
                "field": field,
                "status": "conflict",
                "candidates": conflict_candidates,
                "blocking": field in BLOCKING_CONFLICT_FIELDS,
            })
        field_result = {
            "value": selected["value"],
            "selected_candidate": selected,
            "candidates": items,
            "source": f"{selected.get('source_file')}:{selected.get('source_sheet') or ''}:{selected.get('source_cell') or ''}",
            "confidence": selected.get("confidence"),
        }
        if selected.get("quantity_by_year"):
            field_result.update({
                "unit": "pcs/year",
                "quantity_by_year": selected["quantity_by_year"],
                "qmax": selected.get("qmax"),
                "derivation_method": selected.get("derivation_method"),
            })
        fields[field] = field_result
        resolved_fields.append({"field": field, **field_result})

    missing_fields = [field for field in REQUIRED_COMPONENT_COSTING_FIELDS if not _has_value((fields.get(field) or {}).get("value"))]
    blocking_conflicts = [item for item in conflicts if item["blocking"]]
    return {
        "fields": fields,
        "resolved_fields": sorted(resolved_fields, key=lambda item: item["field"]),
        "missing_fields": missing_fields,
        "conflicts": conflicts,
        "unresolved_conflicts": conflicts,
        "warnings": list(dict.fromkeys(warnings or [])),
        "delivery_location": location,
        "component_costing_ready": not missing_fields and not blocking_conflicts,
    }


def validate_resolved_customer_input(resolution: Dict[str, Any]) -> Dict[str, Any]:
    missing = list(resolution.get("missing_fields") or [])
    conflicts = [item for item in resolution.get("conflicts") or [] if item.get("blocking")]
    ready = not missing and not conflicts
    if ready:
        message = "Customer inputs resolved. Component costing can continue."
    else:
        parts = []
        if missing:
            parts.append("Missing after checking all customer-input files: " + ", ".join(missing))
        if conflicts:
            parts.append("Confirm conflicting fields: " + ", ".join(item["field"] for item in conflicts))
        message = ". ".join(parts) + "."
    return {
        "status": "ready" if ready else "blocked",
        "resolved_fields": resolution.get("resolved_fields") or [],
        "missing_fields": missing,
        "conflicts": conflicts,
        "warnings": resolution.get("warnings") or [],
        "component_costing_ready": ready,
        "message": message,
    }


def extract_customer_input_package(
    structured_payload: Optional[Dict[str, Any]],
    attachments: Any,
) -> Dict[str, Any]:
    discovered = discover_customer_input_files(attachments)
    candidates = []
    warnings = []
    source_reports = []
    green_commercial_inputs = []
    unmapped_green_inputs = []
    structured = extract_structured_request_fields(structured_payload)
    candidates.extend(structured["candidates"])
    source_reports.append(structured)
    for attachment in discovered:
        extension = attachment["extension"]
        path = attachment["path"]
        if extension in {".xlsx", ".xlsm"}:
            report = extract_excel_fields(path)
        elif extension == ".csv":
            report = extract_csv_fields(path)
        elif extension == ".pdf":
            report = extract_pdf_metadata(path)
        else:
            continue
        candidates.extend(report.get("candidates") or [])
        warnings.extend(report.get("warnings") or [])
        source_reports.append(report)
        green_commercial_inputs.extend(report.get("green_commercial_inputs") or [])
        unmapped_green_inputs.extend(report.get("unmapped_green_inputs") or [])
    merged = merge_customer_input_candidates(candidates, warnings)
    validation = validate_resolved_customer_input(merged)
    return {
        **merged,
        "validation": validation,
        "discovered_files": [
            {key: value for key, value in item.items() if key != "path"}
            for item in discovered
        ],
        "source_reports": source_reports,
        "green_commercial_inputs": green_commercial_inputs,
        "unmapped_green_inputs": unmapped_green_inputs,
    }


def apply_resolution_to_customer_input(payload: Dict[str, Any], resolution: Dict[str, Any]) -> Dict[str, Any]:
    result = dict(payload or {})
    values = {field: data.get("value") for field, data in (resolution.get("fields") or {}).items()}
    mapping = {
        "project_code": "project_code",
        "customer": "customer",
        "final_customer": "final_customer",
        "project_name": "project_name",
        "product_name": "product",
        "product_family": "product_line",
        "product_subtype": "product_subtype",
        "part_number": "part_number",
        "annual_quantity": "annual_quantity",
        "sop_date": "sop_date",
        "delivery_country": "delivery_country",
        "delivery_city": "delivery_city",
        "delivery_zone": "customer_delivery_zone",
        "production_plant": "production_plant",
        "target_price": "target_price",
        "target_price_currency": "target_price_currency",
        "quotation_currency": "quotation_currency",
        "purchasing_currency": "purchasing_currency",
        "drawing_reference": "drawing_reference",
        "customer_payment_days": "customer_payment_days",
        "customer_incoterm": "customer_incoterm",
        "customer_delivery_frequency_days": "customer_delivery_frequency_days",
        "platform": "platform",
    }
    for source, destination in mapping.items():
        if _has_value(values.get(source)):
            result[destination] = values[source]
    if _has_value(result.get("part_number")):
        result["part_number"] = str(result["part_number"]).strip()
    result["product_name"] = result.get("product")
    result["currency"] = result.get("quotation_currency")
    annual = resolution.get("fields", {}).get("annual_quantity") or {}
    if annual.get("quantity_by_year"):
        result["quantity_by_year"] = annual["quantity_by_year"]
        result["qmax"] = annual.get("qmax")
        result["annual_quantity_derivation"] = annual.get("derivation_method")
    productivity_percentage = values.get("customer_productivity_percentage")
    productivity_basis = values.get("productivity_basis")
    if _has_value(productivity_percentage) or _has_value(productivity_basis):
        result["customer_productivity"] = {
            **dict(result.get("customer_productivity") or {}),
            **(
                {"percentage": productivity_percentage}
                if _has_value(productivity_percentage) else {}
            ),
            **(
                {"basis": productivity_basis}
                if _has_value(productivity_basis) else {}
            ),
        }
    for field in ("capex_tooling_treatment", "profitability_target", "indexation"):
        if _has_value(values.get(field)):
            result[f"workbook_{field}"] = values[field]
    result["green_commercial_inputs"] = (
        resolution.get("green_commercial_inputs") or []
    )
    result["unmapped_green_inputs"] = (
        resolution.get("unmapped_green_inputs") or []
    )
    return result
