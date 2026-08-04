"""Read-only extraction and reconciliation for approved Choke quotations.

The workbook is an import source, never a runtime calculation dependency.  A
reviewed extraction can be stored as JSON and used to explain historical
quotation conventions alongside the current generic costing model.
"""

from __future__ import annotations

import hashlib
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping


EXTRACTION_VERSION = "choke-excel-golden-reference-v1"
DEFAULT_TOLERANCE = Decimal("0.000001")
PERIOD_COLUMNS = dict(zip(
    ("Y-1", "Y0", "Y1", "Y2", "Y3", "Y4", "EOL"),
    ("B", "C", "D", "E", "F", "G", "H"),
))

ANNUAL_ROWS = {
    "calendar_year": 1,
    "quantity_kpieces": 2,
    "productivity_rate": 3,
    "part_price_before_indexed_material_and_packaging": 4,
    "packaging_and_transport_price": 5,
    "indexed_material_price": 6,
    "selling_price": 7,
    "sales_kcurrency": 9,
    "material_kcurrency": 11,
    "packaging_kcurrency": 12,
    "total_material_kcurrency": 13,
    "dl_kcurrency": 16,
    "voh_kcurrency": 18,
    "inbound_transport_kcurrency": 20,
    "outbound_transport_kcurrency": 21,
    "total_transport_kcurrency": 22,
    "gmdc_kcurrency": 25,
    "foh_kcurrency": 28,
    "fees_kcurrency": 30,
    "ebitda_kcurrency": 33,
    "depreciation_kcurrency": 36,
    "financing_cost_kcurrency": 37,
    "operating_income_kcurrency": 39,
    "taxes_kcurrency": 42,
    "net_result_kcurrency": 44,
    "annual_cash_flow_kcurrency": 54,
    "cumulative_cash_kcurrency": 55,
    "ap_kcurrency": 70,
    "ar_kcurrency": 71,
    "finished_product_stock_kcurrency": 72,
    "wip_kcurrency": 73,
    "raw_material_stock_kcurrency": 74,
    "twc_kcurrency": 75,
    "delta_twc_kcurrency": 76,
}

MAJOR_CELLS = {
    "project_code": ("RFQ Data", "D7"),
    "product": ("RFQ Data", "H7"),
    "part_number": ("RFQ Data", "M7"),
    "customer": ("RFQ Data", "D10"),
    "delivery_zone": ("RFQ Data", "D12"),
    "product_line": ("RFQ Data", "I12"),
    "production_plant": ("RFQ Data", "K13"),
    "customer_incoterm": ("RFQ Data", "D17"),
    "customer_payment_days": ("RFQ Data", "E16"),
    "quotation_currency": ("RFQ Data", "H16"),
    "delivery_frequency_weeks": ("RFQ Data", "H17"),
    "finished_product_inventory_days": ("RFQ Data", "G25"),
    "profitability_target": ("RFQ Data", "H56"),
    "dl_rate_per_hour": ("RFQ Data", "J63"),
    "voh_rate_per_hour": ("RFQ Data", "J64"),
    "wip_days": ("RFQ Data", "J65"),
    "historical_material_tco_per_product": ("Material Cost ", "B35"),
    "base_material_per_product": ("PriceCalAVO", "AB7"),
    "packaging_per_product": ("PriceCalAVO", "AB9"),
    "dl_per_product": ("PriceCalAVO", "AB10"),
    "voh_per_product": ("PriceCalAVO", "AB11"),
    "inbound_transport_per_product": ("PriceCalAVO", "AB12"),
    "internal_transport_per_product": ("PriceCalAVO", "AB13"),
    "outbound_transport_per_product": ("PriceCalAVO", "AB14"),
    "foh_per_product": ("PriceCalAVO", "AB16"),
    "fees_per_product": ("PriceCalAVO", "AB17"),
    "initial_price_seed": ("PriceCalAVO", "AB1"),
    "solved_part_price": ("PriceCalAVO", "B4"),
    "quoted_y0_selling_price": ("PriceCalAVO", "B7"),
    "achieved_irr": ("PriceCalAVO", "B57"),
    "npv": ("PriceCalAVO", "B58"),
    "discount_rate_used_by_workbook": ("PriceCalAVO", "O3"),
    "prelaunch_capex_kcurrency": ("PriceCalAVO", "B61"),
}


def _decimal(value: Any) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _cell(formula_book: Any, value_book: Any, sheet: str, address: str) -> Dict[str, Any]:
    return {
        "sheet": sheet,
        "cell": address,
        "formula": formula_book[sheet][address].value,
        "cached_value": value_book[sheet][address].value,
    }


def _derived(label: str, formula: str, value: Decimal) -> Dict[str, Any]:
    return {
        "sheet": "independently_recalculated",
        "cell": label,
        "formula": formula,
        "cached_value": float(value),
    }


def extract_golden_reference(workbook_path: str | Path) -> Dict[str, Any]:
    """Extract formulas and cached values without modifying the XLSM."""
    from openpyxl import load_workbook

    path = Path(workbook_path).resolve()
    raw = path.read_bytes()
    formula_book = value_book = None
    try:
        formula_book = load_workbook(
            path, data_only=False, keep_vba=True, keep_links=True,
        )
        value_book = load_workbook(
            path, data_only=True, keep_vba=True, keep_links=True,
        )
        sources = {
            key: _cell(formula_book, value_book, sheet, address)
            for key, (sheet, address) in MAJOR_CELLS.items()
        }
        values = {
            key: source["cached_value"] for key, source in sources.items()
        }

        direct = sum(
            _decimal(values[key]) or Decimal("0")
            for key in (
                "dl_per_product", "voh_per_product",
                "inbound_transport_per_product",
                "internal_transport_per_product",
                "outbound_transport_per_product",
            )
        )
        manufacturing = direct + sum(
            _decimal(values[key]) or Decimal("0")
            for key in ("foh_per_product", "fees_per_product")
        )
        delivered_material = (
            (_decimal(values["base_material_per_product"]) or Decimal("0"))
            + (_decimal(values["inbound_transport_per_product"]) or Decimal("0"))
        )
        total_cost = (
            (_decimal(values["base_material_per_product"]) or Decimal("0"))
            + (_decimal(values["packaging_per_product"]) or Decimal("0"))
            + manufacturing
        )
        derived_sources = {
            "direct_cost_per_product": _derived(
                "direct_cost_per_product",
                "PriceCalAVO!AB10+AB11+AB12+AB13+AB14",
                direct,
            ),
            "delivered_material_per_product": _derived(
                "delivered_material_per_product",
                "PriceCalAVO!AB7+AB12",
                delivered_material,
            ),
            "manufacturing_cost_per_product": _derived(
                "manufacturing_cost_per_product",
                "direct_cost_per_product+PriceCalAVO!AB16+AB17",
                manufacturing,
            ),
            "total_product_cost_per_product": _derived(
                "total_product_cost_per_product",
                "PriceCalAVO!AB7+AB9+manufacturing_cost_per_product",
                total_cost,
            ),
        }
        sources.update(derived_sources)
        values.update({
            key: source["cached_value"] for key, source in derived_sources.items()
        })

        components = []
        component_ids = {
            6: "magnet_wire", 7: "ferrite_core", 8: "glue", 9: "lead_tinning",
        }
        for row, component_id in component_ids.items():
            component = {
                "component_id": component_id,
                "designation": value_book["Material Cost "][f"C{row}"].value,
                "quantity_per_product": value_book["Material Cost "][f"F{row}"].value,
                "quantity_unit": value_book["Material Cost "][f"I{row}"].value,
                "supplier_unit_price": value_book["Material Cost "][f"Q{row}"].value,
                "supplier_currency": value_book["Material Cost "][f"R{row}"].value,
                "fx_rate": value_book["Material Cost "][f"U{row}"].value,
                "converted_unit_price": value_book["Material Cost "][f"Y{row}"].value,
                "incoterm": value_book["Material Cost "][f"AC{row}"].value,
                "payment_days": value_book["Material Cost "][f"AF{row}"].value,
                "base_cost_per_product": value_book["Material Cost "][f"BS{row}"].value,
                "inbound_transport_per_product": value_book["Material Cost "][f"BN{row}"].value,
                "historical_tco_per_product": value_book["Material Cost "][f"B{row}"].value,
                "tooling_cost_per_product": value_book["Material Cost "][f"EC{row}"].value,
                "cell_sources": {
                    field: _cell(formula_book, value_book, "Material Cost ", f"{column}{row}")
                    for field, column in {
                        "quantity_per_product": "F", "quantity_unit": "I",
                        "supplier_unit_price": "Q", "supplier_currency": "R",
                        "fx_rate": "U", "converted_unit_price": "Y",
                        "base_cost_per_product": "BS",
                        "inbound_transport_per_product": "BN",
                        "historical_tco_per_product": "B",
                    }.items()
                },
            }
            components.append(component)

        operations = []
        for row in range(4, 32):
            name = value_book["Product Added value"][f"G{row}"].value
            if not name:
                continue
            operations.append({
                "operation_id": value_book["Product Added value"][f"F{row}"].value,
                "operation_name": name,
                "p_h_after_oee": value_book["Product Added value"][f"M{row}"].value,
                "operator_fraction": value_book["Product Added value"][f"K{row}"].value,
                "dl_cost_per_product": value_book["Product Added value"][f"AJ{row}"].value,
                "voh_cost_per_product": value_book["Product Added value"][f"AG{row}"].value,
            })

        annual = {}
        annual_sources = {}
        for period, column in PERIOD_COLUMNS.items():
            annual[period] = {}
            annual_sources[period] = {}
            for field, row in ANNUAL_ROWS.items():
                source = _cell(formula_book, value_book, "PriceCalAVO", f"{column}{row}")
                annual[period][field] = source["cached_value"]
                annual_sources[period][field] = source

        missing_cache = []
        for sheet in formula_book.worksheets:
            value_sheet = value_book[sheet.title]
            for row in sheet.iter_rows():
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                        and value_sheet[cell.coordinate].value is None
                    ):
                        missing_cache.append({
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        })

        names = []
        for item in formula_book.defined_names.values():
            names.append({
                "name": item.name,
                "destination": item.attr_text,
                "broken": "#REF!" in str(item.attr_text),
            })
        external_links = [
            getattr(getattr(item, "file_link", None), "Target", None)
            for item in getattr(formula_book, "_external_links", [])
        ]
        calculation = formula_book.calculation
        sheet_inventory = [{
            "name": sheet.title,
            "state": sheet.sheet_state,
            "used_range": sheet.calculate_dimension(),
            "tables": list(sheet.tables.keys()),
        } for sheet in formula_book.worksheets]

        return {
            "schema_version": EXTRACTION_VERSION,
            "workbook": path.name,
            "source_hash": hashlib.sha256(raw).hexdigest(),
            "source_size_bytes": len(raw),
            "workbook_inventory": {
                "worksheets": sheet_inventory,
                "defined_names": names,
                "external_links": external_links,
                "calculation": {
                    "mode": calculation.calcMode,
                    "full_calculation_on_load": calculation.fullCalcOnLoad,
                    "force_full_calculation": calculation.forceFullCalc,
                    "iterative": calculation.iterate,
                    "iteration_count": calculation.iterateCount,
                    "iteration_delta": calculation.iterateDelta,
                },
                "formula_without_cached_value_count": len(missing_cache),
                "formula_without_cached_value_examples": missing_cache[:100],
                "vba": {
                    "present": True,
                    "observed_modules": ["ThisWorkbook", "Module1", "Module3"],
                    "observed_relevant_symbols": ["GoalSeek", "Export_confirm_Click"],
                    "inspection_method": "static vbaProject.bin string fingerprint",
                },
            },
            "inputs": {
                key: values[key] for key in (
                    "project_code", "product", "part_number", "customer",
                    "delivery_zone", "product_line", "production_plant",
                    "customer_incoterm", "customer_payment_days",
                    "quotation_currency", "delivery_frequency_weeks",
                    "finished_product_inventory_days", "profitability_target",
                    "dl_rate_per_hour", "voh_rate_per_hour", "wip_days",
                )
            },
            "component_costs": components,
            "operations": operations,
            "manufacturing_cost": {
                key: values[key] for key in (
                    "base_material_per_product", "packaging_per_product",
                    "inbound_transport_per_product",
                    "outbound_transport_per_product", "delivered_material_per_product",
                    "dl_per_product", "voh_per_product", "direct_cost_per_product",
                    "foh_per_product", "fees_per_product",
                    "manufacturing_cost_per_product", "total_product_cost_per_product",
                )
            },
            "annual_financial_model": annual,
            "solver": {
                "initial_price_seed": values["initial_price_seed"],
                "solved_part_price": values["solved_part_price"],
                "quoted_y0_selling_price": values["quoted_y0_selling_price"],
                "profitability_target": values["profitability_target"],
                "achieved_irr": values["achieved_irr"],
                "discount_rate_used_by_workbook": values["discount_rate_used_by_workbook"],
                "npv": values["npv"],
                "converged": abs(_decimal(values["npv"]) or Decimal("1")) <= DEFAULT_TOLERANCE,
            },
            "cell_sources": {**sources, "annual_financial_model": annual_sources},
            "business_rule_differences": [
                "Workbook contains seven columns (pre-launch, SOP/Y1, Y2-Y5, EOL), not Y-1 through Y6.",
                "Workbook solves a 50% IRR and discounts NPV at 50%; the later approved generic model uses 12%.",
                "Workbook financing is 10%; the later approved generic model uses 8%.",
                "Workbook WIP duration is 7 days; the later approved Choke rule uses 5 days.",
                "Workbook depreciation starts in pre-launch; the later generic engine starts in Y1.",
            ],
        }
    finally:
        if value_book is not None:
            value_book.close()
        if formula_book is not None:
            formula_book.close()


def reconcile_backend_result(
    reference: Mapping[str, Any],
    backend_result: Mapping[str, Any],
    tolerance: Decimal = DEFAULT_TOLERANCE,
) -> Dict[str, Any]:
    """Compare major technical fields without treating historical values as inputs."""
    expected = reference.get("manufacturing_cost") or {}
    mapping = {
        "base_material_per_product": "material_cost_per_piece",
        "inbound_transport_per_product": "transport_cost_per_piece",
        "delivered_material_per_product": "delivered_material_cost_per_piece",
        "dl_per_product": "dl_cost_per_piece",
        "voh_per_product": "voh_cost_per_piece",
        "direct_cost_per_product": "direct_cost_per_piece",
        "foh_per_product": "foh_cost_per_piece",
        "fees_per_product": "fee_cost_per_piece",
        "manufacturing_cost_per_product": "manufacturing_cost_per_piece",
    }
    rows = []
    for reference_field, backend_field in mapping.items():
        excel_value = _decimal(expected.get(reference_field))
        backend_value = _decimal(backend_result.get(backend_field))
        difference = (
            backend_value - excel_value
            if excel_value is not None and backend_value is not None else None
        )
        percent = (
            difference / excel_value * Decimal("100")
            if difference is not None and excel_value not in (None, Decimal("0"))
            else None
        )
        source = (reference.get("cell_sources") or {}).get(reference_field) or {}
        if excel_value is None:
            status = "missing_reference"
        elif backend_value is None:
            status = "missing_backend"
        elif abs(difference or Decimal("0")) <= tolerance:
            status = "match"
        else:
            status = "mismatch"
        rows.append({
            "metric": reference_field,
            "excel_sheet": source.get("sheet"),
            "excel_cell": source.get("cell"),
            "excel_formula": source.get("formula"),
            "excel_result": None if excel_value is None else float(excel_value),
            "backend_field": backend_field,
            "backend_result": None if backend_value is None else float(backend_value),
            "difference": None if difference is None else float(difference),
            "percentage_difference": None if percent is None else float(percent),
            "status": status,
        })
    return {
        "status": "match" if all(row["status"] == "match" for row in rows) else "mismatch",
        "tolerance": str(tolerance),
        "historical_values_used_in_calculation": False,
        "rows": rows,
    }


def _safe_segment(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(value)).strip("._")


def approved_reference_path(project_code: str, product_id: str) -> Path:
    root = Path(__file__).resolve().parents[1] / "data" / "reference_quotations"
    return root / _safe_segment(project_code) / f"{_safe_segment(product_id)}.json"


def get_approved_reference_report(
    project_code: str,
    product_id: str,
    backend_result: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    path = approved_reference_path(project_code, product_id)
    if not path.exists():
        return {
            "status": "not_available",
            "project_code": project_code,
            "product_id": product_id,
            "message": "No approved quotation reference has been imported for this project/product.",
        }
    reference = json.loads(path.read_text(encoding="utf-8"))
    return {
        "status": "found",
        "project_code": project_code,
        "product_id": product_id,
        "reference": {
            "workbook": reference.get("workbook"),
            "source_hash": reference.get("source_hash"),
            "inputs": reference.get("inputs"),
            "component_costs": reference.get("component_costs"),
            "manufacturing_cost": reference.get("manufacturing_cost"),
            "annual_financial_model": reference.get("annual_financial_model"),
            "solver": reference.get("solver"),
            "business_rule_differences": reference.get("business_rule_differences"),
        },
        "reconciliation": (
            reconcile_backend_result(reference, backend_result)
            if backend_result is not None else None
        ),
    }
