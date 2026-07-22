import os
import sys
from pathlib import Path


ROOT_DIR = Path(__file__).resolve().parents[1]
TARGET_FILE = "26185-CHO-TONGKAI-ROD Chokes KT-317-China-Assembly Quotation(2).xlsm"
KEYWORDS = [
    "Material",
    "Transport",
    "DL",
    "VOH",
    "FOH",
    "FEE",
    "Price",
    "Tooling",
    "CAPEX",
    "VAN",
    "ROI",
]


def _candidate_roots():
    roots = [
        ROOT_DIR,
        ROOT_DIR.parent,
        ROOT_DIR / "data",
        ROOT_DIR.parent / "data",
        Path.home() / "Downloads",
    ]
    configured = os.getenv("PRICECALAVO_WORKBOOK_PATH")
    if configured:
        path = Path(configured)
        roots.insert(0, path if path.is_dir() else path.parent)
    return list(dict.fromkeys(root for root in roots if root.exists()))


def find_workbook():
    configured = os.getenv("PRICECALAVO_WORKBOOK_PATH")
    if configured:
        path = Path(configured)
        if not path.is_absolute():
            path = ROOT_DIR / path
        if path.exists():
            return path

    for root in _candidate_roots():
        direct = root / TARGET_FILE
        if direct.exists():
            return direct
        try:
            matches = list(root.rglob(TARGET_FILE))
        except Exception:
            matches = []
        if matches:
            return matches[0]
    return None


def _cell_display(value):
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    if len(text) > 120:
        return text[:117] + "..."
    return text


def _inspect_loaded_workbook(workbook):
    print()
    print("Sheet names:")
    for sheet_name in workbook.sheetnames:
        print(f"- {sheet_name}")

    if "PriceCalAVO" not in workbook.sheetnames:
        print()
        print("Sheet PriceCalAVO not found.")
        return 0

    sheet = workbook["PriceCalAVO"]
    print()
    print("PriceCalAVO used range:")
    print(f"- rows: {sheet.max_row}")
    print(f"- columns: {sheet.max_column}")

    max_rows = min(sheet.max_row, 80)
    max_cols = min(sheet.max_column, 30)
    print()
    print(f"First {max_rows} rows x {max_cols} columns values/formulas:")
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
        values = [_cell_display(cell.value) for cell in row]
        print(f"R{row[0].row:03d}: " + " | ".join(values))

    print()
    print("Keyword matches:")
    lower_keywords = [(keyword, keyword.lower()) for keyword in KEYWORDS]
    matches = []
    for row in sheet.iter_rows():
        for cell in row:
            value = _cell_display(cell.value)
            if not value:
                continue
            lower_value = value.lower()
            for keyword, lower_keyword in lower_keywords:
                if lower_keyword in lower_value:
                    matches.append((keyword, cell.coordinate, cell.row, cell.column, value))

    if not matches:
        print("- No keyword matches found.")
        return 0

    for keyword, coordinate, row_number, column_number, value in matches[:300]:
        print(f"- {keyword}: {coordinate} (row {row_number}, col {column_number}) = {value}")
    if len(matches) > 300:
        print(f"- Truncated {len(matches) - 300} additional matches.")
    return 0


def _close_workbook(workbook):
    archive = getattr(workbook, "_archive", None)
    vba_archive = getattr(workbook, "vba_archive", None)
    try:
        workbook.close()
    finally:
        if vba_archive is not None and vba_archive is not archive:
            vba_archive.close()


def main():
    workbook_path = find_workbook()
    print("PRICECALAVO WORKBOOK INSPECTION")
    print("=" * 78)
    print(f"Target workbook: {TARGET_FILE}")
    if not workbook_path:
        print("Workbook not found.")
        print("Set PRICECALAVO_WORKBOOK_PATH or place the file under mcp_server/data or Downloads.")
        return 0

    print(f"Workbook path: {workbook_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"openpyxl is unavailable, cannot inspect workbook: {exc}")
        return 0

    try:
        workbook = load_workbook(workbook_path, read_only=False, keep_vba=True, data_only=False)
    except Exception as exc:
        print(f"Could not open workbook: {exc}")
        return 0

    try:
        return _inspect_loaded_workbook(workbook)
    finally:
        _close_workbook(workbook)


if __name__ == "__main__":
    sys.exit(main())
